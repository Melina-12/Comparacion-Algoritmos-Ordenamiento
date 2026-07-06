# ==========================================
# ANÁLISIS COMPARATIVO DE ALGORITMOS
# Bubble Sort - Selection Sort - Quick Sort
# Python 3.13
# ==========================================

import random
import time
import statistics
import sys
import os

import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Permite trabajar con Quick Sort en listas grandes
sys.setrecursionlimit(50000)

# Crear carpeta para guardar gráficos
os.makedirs("graficos", exist_ok=True)
# ==========================================
# PARÁMETROS DEL EXPERIMENTO
# ==========================================

# Tamaños de prueba
TAMANOS = [
    100,
    500,
    1000,
    2500,
    5000,
    7500,
    10000
]

# Número de repeticiones
REPETICIONES = 10
# ==========================================
# BUBBLE SORT
# ==========================================

def bubble_sort(lista):

    datos = lista.copy()

    n = len(datos)

    for i in range(n):

        intercambio = False

        # Recorre el arreglo comparando elementos vecinos
        for j in range(n - i - 1):

            if datos[j] > datos[j + 1]:

                datos[j], datos[j + 1] = datos[j + 1], datos[j]

                intercambio = True

        # Si no hubo cambios, el arreglo ya está ordenado
        if not intercambio:
            break

    return datos

# ==========================================
# SELECTION SORT
# ==========================================

def selection_sort(lista):

    datos = lista.copy()

    n = len(datos)

    for i in range(n):

        minimo = i

        # Busca el menor elemento restante
        for j in range(i + 1, n):

            if datos[j] < datos[minimo]:

                minimo = j

        datos[i], datos[minimo] = datos[minimo], datos[i]

    return datos

# ==========================================
# QUICK SORT
# ==========================================

def quick_sort(lista):

    if len(lista) <= 1:
        return lista

    pivote = lista[len(lista)//2]

    menores = [x for x in lista if x < pivote]

    iguales = [x for x in lista if x == pivote]

    mayores = [x for x in lista if x > pivote]

    return quick_sort(menores) + iguales + quick_sort(mayores)

# ==========================================
# VERIFICAR ALGORITMOS
# ==========================================

def verificar_algoritmos():

    prueba = [5, 8, 1, 6, 9, 2, 4, 7, 3]

    correcto = sorted(prueba)

    assert bubble_sort(prueba) == correcto

    assert selection_sort(prueba) == correcto

    assert quick_sort(prueba) == correcto

    print("✓ Todos los algoritmos funcionan correctamente.\n")
# ==========================================
# MEDICIÓN DEL TIEMPO
# ==========================================

def medir_tiempo(funcion, datos):

    inicio = time.perf_counter()

    funcion(datos)

    fin = time.perf_counter()

    return fin - inicio


# ==========================================
# GENERAR DATOS
# ==========================================

def generar_datos(tamano, escenario):

    if escenario == "Aleatorio":
        return [random.randint(1, 100000) for _ in range(tamano)]

    elif escenario == "Ordenado":
        return list(range(tamano))

    elif escenario == "Inverso":
        return list(range(tamano, 0, -1))

    else:
        raise ValueError("Escenario no válido")
    
# ==========================================
# EJECUTAR EXPERIMENTO
# ==========================================

def ejecutar_experimento():

    resultados = []

    escenarios = [
        "Aleatorio",
        "Ordenado",
        "Inverso"
    ]

    print("\nIniciando pruebas experimentales...\n")

    for escenario in escenarios:

        print(f"Escenario: {escenario}")

        for tamano in TAMANOS:

            tiempos_bubble = []
            tiempos_selection = []
            tiempos_quick = []

            for _ in range(REPETICIONES):

                datos = generar_datos(tamano, escenario)

                tiempos_bubble.append(
                    medir_tiempo(bubble_sort, datos)
                )

                tiempos_selection.append(
                    medir_tiempo(selection_sort, datos)
                )

                tiempos_quick.append(
                    medir_tiempo(quick_sort, datos)
                )

            resultados.append({

                "Escenario": escenario,

                "Tamaño": tamano,

                "Bubble Promedio":
                    statistics.mean(tiempos_bubble),

                "Bubble Mediana":

                    statistics.median(tiempos_bubble),

                "Bubble Desv":
                    statistics.stdev(tiempos_bubble),

                "Selection Promedio":
                    statistics.mean(tiempos_selection),
                "Selection Mediana":
                    statistics.median(tiempos_selection),

                "Selection Desv":
                    statistics.stdev(tiempos_selection),

                "Quick Promedio":
                    statistics.mean(tiempos_quick),
                "Quick Mediana":
                    statistics.median(tiempos_quick),

                "Quick Desv":
                    statistics.stdev(tiempos_quick)

            })

            print(f"   Tamaño {tamano} completado.")

        print()

    return pd.DataFrame(resultados)


# ==========================================
# GUARDAR RESULTADOS
# ==========================================

def guardar_resultados(tabla):

    with pd.ExcelWriter("resultados.xlsx", engine="openpyxl") as writer:

        tabla.to_excel(writer,
                       sheet_name="Resultados completos",
                       index=False)

        tabla[
            ["Escenario","Tamaño",
             "Bubble Promedio","Selection Promedio","Quick Promedio"]
        ].to_excel(writer,
                   sheet_name="Promedios",
                   index=False)

        tabla[
            ["Escenario","Tamaño",
             "Bubble Mediana","Selection Mediana","Quick Mediana"]
        ].to_excel(writer,
                   sheet_name="Medianas",
                   index=False)

        tabla[
            ["Escenario","Tamaño",
             "Bubble Desv","Selection Desv","Quick Desv"]
        ].to_excel(writer,
                   sheet_name="Desviaciones",
                   index=False)

        # Dar formato a todas las hojas
        for hoja in writer.sheets.values():

            # Encabezados
            for celda in hoja[1]:
                celda.font = Font(bold=True, color="FFFFFF")
                celda.fill = PatternFill(
                    fill_type="solid",
                    start_color="1F4E78"
                )
                celda.alignment = Alignment(horizontal="center")

            # Ajustar ancho de columnas
            for columna in hoja.columns:

                longitud = max(len(str(celda.value))
                               if celda.value is not None else 0
                               for celda in columna)

                letra = get_column_letter(columna[0].column)

                hoja.column_dimensions[letra].width = longitud + 3

            # Formato decimal
            for fila in hoja.iter_rows(min_row=2):
                for celda in fila:
                    if isinstance(celda.value, float):
                        celda.number_format = "0.000000"
                        tabla.to_csv(
    "resultados.csv",
    sep=";",
    index=False,
    encoding="utf-8-sig"
)

    print("\n✓ Archivo Excel generado correctamente.")
# ==========================================
# GENERAR GRÁFICOS
# ==========================================

def generar_graficos(tabla):

    escenarios = tabla["Escenario"].unique()

    for escenario in escenarios:

        datos = tabla[tabla["Escenario"] == escenario]

        plt.figure(figsize=(10,6))

        plt.errorbar(
            datos["Tamaño"],
            datos["Bubble Promedio"],
            yerr=datos["Bubble Desv"],
            marker="o",
            linewidth=2,
            capsize=5,
            label="Bubble Sort"
        )

        plt.errorbar(
            datos["Tamaño"],
            datos["Selection Promedio"],
            yerr=datos["Selection Desv"],
            marker="s",
            linewidth=2,
            capsize=5,
            label="Selection Sort"
        )

        plt.errorbar(
            datos["Tamaño"],
            datos["Quick Promedio"],
            yerr=datos["Quick Desv"],
            marker="^",
            linewidth=2,
            capsize=5,
            label="Quick Sort"
        )

        plt.title(f"Comparación de tiempos - {escenario}")

        plt.xlabel("Número de elementos")

        plt.ylabel("Tiempo promedio (segundos)")

        plt.yscale("log")

        plt.grid(True)

        plt.legend()

        plt.tight_layout()

        plt.savefig(
            f"graficos/{escenario}.png",
            dpi=300,
            bbox_inches="tight"
        )

        plt.close()

    print("\n✓ Gráficos individuales generados.")

# ==========================================
# RESUMEN ESTADÍSTICO
# ==========================================

def resumen_estadistico(tabla):

    print("\n===================================")
    print("RESUMEN DEL EXPERIMENTO")
    print("===================================\n")

    escenarios = tabla["Escenario"].unique()

    for escenario in escenarios:

        datos = tabla[tabla["Escenario"] == escenario]

        promedio_bubble = datos["Bubble Promedio"].mean()
        promedio_selection = datos["Selection Promedio"].mean()
        promedio_quick = datos["Quick Promedio"].mean()

        tiempos = {
            "Bubble Sort": promedio_bubble,
            "Selection Sort": promedio_selection,
            "Quick Sort": promedio_quick
        }

        mejor = min(tiempos, key=tiempos.get)

        print(f"Escenario: {escenario}")

        print(f"Bubble Sort    : {promedio_bubble:.6f} s")
        print(f"Selection Sort : {promedio_selection:.6f} s")
        print(f"Quick Sort     : {promedio_quick:.6f} s")

        print(f"Algoritmo más rápido: {mejor}\n")
# ==========================================
# GRÁFICO GENERAL
# ==========================================

def grafico_general(tabla):

    promedio = tabla.groupby("Tamaño").agg({

        "Bubble Promedio":"mean",
        "Bubble Desv":"mean",

        "Selection Promedio":"mean",
        "Selection Desv":"mean",

        "Quick Promedio":"mean",
        "Quick Desv":"mean"

    })

    plt.figure(figsize=(11,6))

    plt.errorbar(
        promedio.index,
        promedio["Bubble Promedio"],
        yerr=promedio["Bubble Desv"],
        marker="o",
        linewidth=2,
        capsize=5,
        label="Bubble Sort"
    )

    plt.errorbar(
        promedio.index,
        promedio["Selection Promedio"],
        yerr=promedio["Selection Desv"],
        marker="s",
        linewidth=2,
        capsize=5,
        label="Selection Sort"
    )

    plt.errorbar(
        promedio.index,
        promedio["Quick Promedio"],
        yerr=promedio["Quick Desv"],
        marker="^",
        linewidth=2,
        capsize=5,
        label="Quick Sort"
    )

    plt.title("Comparación General de Algoritmos")

    plt.xlabel("Número de elementos")

    plt.ylabel("Tiempo promedio (segundos)")

    plt.yscale("log")

    plt.grid(True, which="both", linestyle="--")

    plt.legend()

    plt.tight_layout()

    plt.savefig(
        "graficos/Comparacion_General.png",
        dpi=300,
        bbox_inches="tight"
    )

    plt.close()

    print("✓ Gráfico general generado.")

if __name__ == "__main__":

    verificar_algoritmos()

    tabla = ejecutar_experimento()

    print("\nResultados obtenidos:\n")

    print(tabla)

    guardar_resultados(tabla)

    generar_graficos(tabla)

    grafico_general(tabla)

    resumen_estadistico(tabla)

    print("\nProyecto finalizado correctamente.") 